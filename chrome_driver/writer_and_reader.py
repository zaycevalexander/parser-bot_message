from openpyxl import load_workbook


def write_url(dictionary):
    wb = load_workbook('./file.xlsx')
    sheet = wb.active
    count = dictionary['count']
    for i in dictionary['list_value']:
        sheet[f'A{count}'] = i
        count += 1
        wb.save('./file.xlsx')
        wb.close()


def read_file():
    list_value = []
    wb = load_workbook('./file.xlsx')
    sheet = wb.active
    count = sheet.max_row + 1
    for i in range(1, count):
        list_value.append(sheet.cell(row=i, column=1).value)
    wb.close()
    dict_params = {
        'list_value': list_value,
        'count': count
    }
    return dict_params
